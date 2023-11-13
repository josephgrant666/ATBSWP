import openpyxl, os

os.chdir('where you download the spreadsheet file')
workbook = openpyxl.load_workbook('relative file path') # The function that is used to open spreadsheets 
sheet = workbook.get_sheet_by_name('Sheet1')            # The function that is used to locate the name of a sheet inside of a spreadsheet
workbook.get_sheet_names()                              # Returns all sheet names 
sheet['A1']                                             # Returns the location of cell A1
cell.value                                              # Returns the values found in this cell
str(sheet['A1'].value)                                  # This returns a neat string of what is in that cell
sheet.cell(row=1,column=2)                              # Returns the object found at these coordinates
for i in range(1, 8):
    print(i, sheet.cell(row=i,column=2).value)          # Returns the values found in each row according to i
